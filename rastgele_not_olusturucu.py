#Gerekli kütüphanelerin aktif hale getirilmesi
from random import choice, randint, random
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

wb = Workbook()
ws = wb.active

hn = {
20: "A1", 19:"A1", 18:"A2", 17:"A3", 16:"B1", 15:"B2", 14:"B3", 13:"C1", 12:"C2", 11:"C3", 10:"D", 9:"F3", 8:"F3" , 7:"F3", 6:"F3", 5:"F3", 4:"F3", 3:"F3", 2:"F3", 1:"F3", 0:"F3"
}

db = Faker('tr')


def change(n) -> int:
    """
    İki sınav notu arasındaki farkı hesaplar. 
    """
    c = (random() * 2) - 1
    return int(n*c)

def update(x, degisim, bonus=0) -> int:
    """
    İlk notu kullanarak bir sonraki rastgele puanları hesaplar.
    """
    result = x + change(degisim) + bonus
    if result < 0:
        result += 5
        return update(result, degisim)
    elif result > 100:
        result -= 5
        return update(result, degisim)
    else:
        return result


#Kalın ve ince kenarlıkların oluşturulması:
thick_border = Border(left=Side(style="thick"), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thin_border = Border(left=Side(style="thin"), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thickBorder_upDown = Border(left=Side(style="dotted"), 
                            right=Side(style="dotted"),
                            top=Side(style="thick"),
                            bottom=Side(style="thick"))

#Başlıkların excel dosyasına eklenmesi:
ws.append(['İsim', 'Vize-1', 'Vize-2', 'Final',  "Ortalama" ,'Harf Notu', "Geçme Durumu"])

#Sınavaların ortalaması hesaplanırken bu değişkenler kullanılacak. 
_v1 , _v2, _f , _o = 0, 0, 0, 0
numberOfStudents = 100

for i in range(numberOfStudents):
    name = db.first_name() #Rastgele bir isim oluşturuluyor
    v1 = randint(0,100)
    _v1 += v1
    bonus = 0
    if v1 < 50:
        bonus += 10
    v2 = update(v1, 30, bonus)
    _v2 += v2
    if v2 > 85:
        bonus -=5
    if (v1+v2) < 100:
        bonus += 10
    if (v1 + v2) < 40:
        bonus -= 10
    f = update(v2,30, 10)
    if (v1 + v2) <=15:
        f = choice([0,1,0,1,0,2,3,0,4,0,5])
    _f += f
    ortalama = (v1*0.3 + v2*0.3 + f*0.4)
    ortalama = round(ortalama, 2)
    _o += ortalama
    harfNotu = hn[ortalama//5] #Ortalamanın 5'e tam bölünmesi ile harf notu elde ediliyor. 
    gecmeDurumu = "Geçti"
    if harfNotu == "F3":
        gecmeDurumu = "Kaldı"
    L = [name,v1, v2, f, ortalama, harfNotu, gecmeDurumu] #Yeni satıra eklenecek elemanların listesi
    ws.append(L)


#Son satırın eklenmesi (Ortalamalar)
ws.append(("Genel Ortalama: ", (_v1/numberOfStudents), (_v2/numberOfStudents), (_f/numberOfStudents), (_o/numberOfStudents)))


#Görsel düzenlemeler:

#Notların girili olduğu hücrelerin hizalanması ve kalın kenarlıklar:
range = ws[f'B1:F{numberOfStudents+2}']
for row in range:
    for cell in row:
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

range = ws[f"B1:E{numberOfStudents+2}"]
for row in range:
    for cell in row:
        cell.border = thickBorder_upDown

range = ws['A']
for cell in range:
    cell.border = thick_border

ws.column_dimensions["A"].width = 15

#Harf notlarının renklerine özgü olarak boyanması:
harfNotuColumn = ws[f"F1:F{numberOfStudents+1}"]
aPaint = PatternFill("solid",start_color="00FF00", end_color="00FF00")
bPaint = PatternFill("solid",start_color="CCCCFF", end_color="CCCCFF")
cPaint = PatternFill("solid",start_color="FF8080", end_color="FF8080")
dPaint = PatternFill("solid",start_color="008000", end_color="008000")
fPaint = PatternFill("solid",start_color="FF0000", end_color="FF0000")

boldFont = Font(name="calibri", bold=True, size=11)
for cell in harfNotuColumn:
    cell = cell[0]
    cell.border = thick_border
    if 'A' in cell.value:   cell.fill = aPaint
    elif 'B' in cell.value: cell.fill = bPaint
    elif 'C' in cell.value: cell.fill = cPaint
    elif 'D' in cell.value: cell.fill = dPaint
    elif 'F' in cell.value: cell.fill = fPaint

    #Harf notlarının simgesini kalınlaştırıyoruz
    cell.font = boldFont
    

#Geçti ve Kaldı bilgisinin yazılı olduğu sütundaki kareleri içlerindeki yazıya göre boyuyoruz:
gectiKaldiColumn = ws[f"G1:G{numberOfStudents+1}"]
for _cell in gectiKaldiColumn:
    cell = _cell[0]
    cell.font = boldFont
    cell.border = thick_border
    if cell.value == "Kaldı": cell.fill = PatternFill("solid", start_color="00FFFF", end_color="00FFFF")
    else:                     cell.fill = PatternFill("solid", start_color="00FF00", end_color="00FF00")

wb.save("liste.xlsx") #Dosyayı kaydediyoruz 
os.system("liste.xlsx") #Dosyayı çalıştırıyoruz